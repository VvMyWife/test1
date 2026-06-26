#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import argparse
import datetime as dt
import os
import re
import shlex
import subprocess
import sys
from pathlib import Path, PurePosixPath

from openpyxl import load_workbook


DEFAULT_PROJECT_DIR = "/home/ubuntu/mineru_workspace"

DEFAULT_XLSX = (
    "/home/ubuntu/mineru_workspace/data/input/"
    "OCR0706刘家诚/第一批自测数据/测试数据/反标测试点.xlsx"
)

DEFAULT_INPUT_HOST_DIR = (
    "/home/ubuntu/mineru_workspace/data/input/"
    "OCR0706刘家诚/第一批自测数据/测试数据"
)

DEFAULT_INPUT_CONTAINER_DIR = (
    "/workspace/input/OCR0706刘家诚/第一批自测数据/测试数据"
)

DEFAULT_OUTPUT_CONTAINER_DIR = (
    "/workspace/output/OCR0706刘家诚/第一批自测数据/测试数据"
)

DEFAULT_SERVICE = "mineru-operator"
DEFAULT_TABLE_ENGINE = "paddle"
DEFAULT_CONCURRENCY = 8

DEFAULT_EXTENSIONS = [
    ".jpg",
    ".jpeg",
    ".png",
    ".pdf",
    ".tif",
    ".tiff",
    ".bmp",
]


def normalize_text(value):
    if value is None:
        return ""

    if isinstance(value, dt.datetime):
        if value.time() == dt.time(0, 0, 0):
            return value.strftime("%Y-%m-%d")
        return value.strftime("%Y-%m-%d %H:%M:%S")

    if isinstance(value, dt.date):
        return value.strftime("%Y-%m-%d")

    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return str(value).strip()

    text = str(value).strip()
    text = re.sub(r"\s+", " ", text)
    return text


def normalize_file_no(value, width):
    text = normalize_text(value)

    if not text:
        return ""

    path = Path(text)
    if path.suffix:
        return text

    if re.fullmatch(r"\d+", text):
        return text.zfill(width)

    return text


def load_rows_from_xlsx(xlsx_path):
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb.active

    header_row = 1
    headers = {}

    for cell in ws[header_row]:
        name = normalize_text(cell.value)
        if name:
            headers[name] = cell.column

    if "文件号" not in headers:
        raise RuntimeError("Excel 表中没有找到表头：文件号")

    keyword_headers = []
    for name, col_idx in headers.items():
        if name.startswith("测试点"):
            keyword_headers.append((name, col_idx))

    keyword_headers.sort(key=lambda x: x[1])

    if not keyword_headers:
        raise RuntimeError("Excel 表中没有找到测试点列，例如：测试点1、测试点2、测试点3")

    rows = []

    for row_idx in range(header_row + 1, ws.max_row + 1):
        file_cell = ws.cell(row=row_idx, column=headers["文件号"])
        raw_file_no = file_cell.value

        if raw_file_no is None:
            continue

        keywords = []
        for _, col_idx in keyword_headers:
            value = ws.cell(row=row_idx, column=col_idx).value
            text = normalize_text(value)
            if text:
                keywords.append(text)

        rows.append(
            {
                "excel_row": row_idx,
                "raw_file_no": raw_file_no,
                "keywords": keywords,
            }
        )

    return rows


def find_input_file(input_host_dir, file_name_or_no, extensions):
    input_host_dir = Path(input_host_dir)

    candidate_text = str(file_name_or_no).strip()
    candidate_path = Path(candidate_text)

    if candidate_path.suffix:
        direct = input_host_dir / candidate_text
        if direct.exists() and direct.is_file():
            return direct

        matches = list(input_host_dir.rglob(candidate_text))
        matches = [p for p in matches if p.is_file()]
        if matches:
            return sorted(matches)[0]

        return None

    for ext in extensions:
        direct = input_host_dir / f"{candidate_text}{ext}"
        if direct.exists() and direct.is_file():
            return direct

    for ext in extensions:
        matches = list(input_host_dir.rglob(f"{candidate_text}{ext}"))
        matches = [p for p in matches if p.is_file()]
        if matches:
            return sorted(matches)[0]

    all_matches = list(input_host_dir.rglob(f"{candidate_text}.*"))
    all_matches = [p for p in all_matches if p.is_file()]
    if all_matches:
        return sorted(all_matches)[0]

    return None


def to_container_path(host_file, input_host_dir, input_container_dir):
    host_file = Path(host_file).resolve()
    input_host_dir = Path(input_host_dir).resolve()

    rel_path = host_file.relative_to(input_host_dir)
    container_path = PurePosixPath(input_container_dir) / PurePosixPath(rel_path.as_posix())
    return str(container_path)


def build_command(args, container_file, keywords):
    cmd = [
        "docker",
        "compose",
        "exec",
        "-T",
        args.service,
        "mineru-operator-batch",
        container_file,
        "--output-dir",
        args.output_container_dir,
        "--table-engine",
        args.table_engine,
    ]

    if keywords:
        cmd.extend(
            [
                "--field-keywords",
                args.keyword_joiner.join(keywords),
            ]
        )

    cmd.extend(
        [
            "--concurrency",
            str(args.concurrency),
            "--overwrite",
            "--recursive",
        ]
    )

    return cmd


def shell_join(cmd):
    return " ".join(shlex.quote(str(part)) for part in cmd)


def ensure_log_dir(project_dir):
    log_dir = Path(project_dir) / "logs"
    log_dir.mkdir(parents=True, exist_ok=True)
    timestamp = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
    return log_dir / f"field_keywords_batch_{timestamp}.log"


def write_log(log_file, text):
    with open(log_file, "a", encoding="utf-8") as f:
        f.write(text)
        if not text.endswith("\n"):
            f.write("\n")


def parse_args():
    parser = argparse.ArgumentParser(
        description="Run mineru-operator-batch by reading file names and field keywords from an Excel file."
    )

    parser.add_argument("--project-dir", default=DEFAULT_PROJECT_DIR)
    parser.add_argument("--xlsx", default=DEFAULT_XLSX)
    parser.add_argument("--input-host-dir", default=DEFAULT_INPUT_HOST_DIR)
    parser.add_argument("--input-container-dir", default=DEFAULT_INPUT_CONTAINER_DIR)
    parser.add_argument("--output-container-dir", default=DEFAULT_OUTPUT_CONTAINER_DIR)

    parser.add_argument("--service", default=DEFAULT_SERVICE)
    parser.add_argument("--table-engine", default=DEFAULT_TABLE_ENGINE)
    parser.add_argument("--concurrency", type=int, default=DEFAULT_CONCURRENCY)

    parser.add_argument("--file-no-width", type=int, default=4)
    parser.add_argument(
        "--extensions",
        default=",".join(DEFAULT_EXTENSIONS),
        help="Comma-separated file extensions used when resolving file numbers.",
    )

    parser.add_argument(
        "--keyword-mode",
        choices=["joined"],
        default="joined",
        help="joined means '--field-keywords kw1,kw2,kw3'.",
    )
    parser.add_argument("--keyword-joiner", default=",")

    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument("--limit", type=int, default=0)
    parser.add_argument("--start-index", type=int, default=1)
    parser.add_argument("--stop-on-error", action="store_true")

    return parser.parse_args()


def main():
    args = parse_args()

    project_dir = Path(args.project_dir)
    xlsx_path = Path(args.xlsx)
    input_host_dir = Path(args.input_host_dir)

    if not project_dir.exists():
        raise RuntimeError(f"Project dir does not exist: {project_dir}")

    if not xlsx_path.exists():
        raise RuntimeError(f"Excel file does not exist: {xlsx_path}")

    if not input_host_dir.exists():
        raise RuntimeError(f"Input host dir does not exist: {input_host_dir}")

    os.chdir(project_dir)

    extensions = []
    for item in args.extensions.split(","):
        item = item.strip()
        if not item:
            continue
        if not item.startswith("."):
            item = "." + item
        extensions.append(item.lower())

    rows = load_rows_from_xlsx(xlsx_path)

    if args.start_index > 1:
        rows = rows[args.start_index - 1 :]

    if args.limit > 0:
        rows = rows[: args.limit]

    log_file = ensure_log_dir(project_dir)

    total = len(rows)
    success_count = 0
    failed_items = []
    skipped_items = []

    write_log(log_file, f"Project dir: {project_dir}")
    write_log(log_file, f"Excel file: {xlsx_path}")
    write_log(log_file, f"Input host dir: {input_host_dir}")
    write_log(log_file, f"Input container dir: {args.input_container_dir}")
    write_log(log_file, f"Output container dir: {args.output_container_dir}")
    write_log(log_file, f"Total rows: {total}")
    write_log(log_file, "")

    print(f"共读取到 {total} 条任务")
    print(f"日志文件：{log_file}")

    for idx, row in enumerate(rows, start=args.start_index):
        file_no = normalize_file_no(row["raw_file_no"], args.file_no_width)
        keywords = row["keywords"]

        if not file_no:
            skipped_items.append((idx, row["excel_row"], "empty file no"))
            print(f"[SKIP] index={idx}, excel_row={row['excel_row']}，文件号为空")
            continue

        if not keywords:
            skipped_items.append((idx, row["excel_row"], f"{file_no}: empty keywords"))
            print(f"[SKIP] index={idx}, excel_row={row['excel_row']}，{file_no} 没有测试点")
            continue

        host_file = find_input_file(input_host_dir, file_no, extensions)

        if host_file is None:
            failed_items.append((idx, row["excel_row"], file_no, "input file not found"))
            print(f"[MISS] index={idx}, excel_row={row['excel_row']}，找不到文件：{file_no}")
            write_log(log_file, f"[MISS] index={idx}, excel_row={row['excel_row']}, file={file_no}")
            continue

        container_file = to_container_path(
            host_file=host_file,
            input_host_dir=input_host_dir,
            input_container_dir=args.input_container_dir,
        )

        cmd = build_command(args, container_file, keywords)
        cmd_text = shell_join(cmd)

        print("=" * 100)
        print(f"[{idx}/{args.start_index + total - 1}] Excel行={row['excel_row']} 文件={host_file.name}")
        print(f"测试点：{keywords}")
        print(cmd_text)

        write_log(log_file, "=" * 100)
        write_log(log_file, f"[TASK] index={idx}, excel_row={row['excel_row']}, host_file={host_file}")
        write_log(log_file, f"[KEYWORDS] {keywords}")
        write_log(log_file, f"[COMMAND] {cmd_text}")

        if args.dry_run:
            continue

        result = subprocess.run(
            cmd,
            cwd=str(project_dir),
            text=True,
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
        )

        write_log(log_file, "[STDOUT]")
        write_log(log_file, result.stdout or "")
        write_log(log_file, "[STDERR]")
        write_log(log_file, result.stderr or "")
        write_log(log_file, f"[RETURN_CODE] {result.returncode}")

        if result.returncode == 0:
            success_count += 1
            print(f"[OK] {host_file.name}")
        else:
            failed_items.append((idx, row["excel_row"], host_file.name, f"return code {result.returncode}"))
            print(f"[FAIL] {host_file.name}，return code={result.returncode}")

            if args.stop_on_error:
                print("因为启用了 --stop-on-error，脚本已停止。")
                break

    print("=" * 100)

    if args.dry_run:
        print("当前是 dry-run 模式，只打印命令，没有真正执行。")
    else:
        print(f"执行完成：成功 {success_count} 条，失败 {len(failed_items)} 条，跳过 {len(skipped_items)} 条。")

    if failed_items:
        print("\n失败列表：")
        for item in failed_items:
            print(f"  index={item[0]}, excel_row={item[1]}, file={item[2]}, reason={item[3]}")

    if skipped_items:
        print("\n跳过列表：")
        for item in skipped_items:
            print(f"  index={item[0]}, excel_row={item[1]}, reason={item[2]}")

    print(f"\n详细日志：{log_file}")


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("\n用户中断。", file=sys.stderr)
        sys.exit(130)
    except Exception as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        sys.exit(1)